
from datetime import datetime
import sys
from PyQt5 import QtWidgets
from PyQt5 import QtCore, QtGui

from easygui import fileopenbox
from easygui import diropenbox
import platform
import xml.etree.ElementTree as ET
import functools
import json
from openpyxl import load_workbook
import os

from MainWindow import Ui_MainWindow_NMR
from GUI_Toolbox import TableModelData
from GUI_Toolbox import ComparisonTableModelData
from GUI_Toolbox import NMRData
from Driver_referenceMeasurement_createFile_importExcel import Driver_referenceMeasurement_createFile_importExcel
from Driver_surfaceAreaCalculation_oneRelaxationTime_createFile_importExcel import Driver_surfaceAreaCalculation_oneRelaxationTime_createFile_importExcel
from Driver_comparisonPlots import Driver_comparisonPlots
from Driver_referenceMeasurement_createFile_withoutExcel import Driver_referenceMeasurement_createFile_withoutExcel

import pandas as pd




class GUI_MainWindow:
    
    #Get UI components

    
    
    def __init__(self):

        self.main_win = QtWidgets.QMainWindow()
        self.ui = Ui_MainWindow_NMR()
        self.ui.setupUi(self.main_win)


        #READ JSON DATABASE OF MATERIAL/BULK
        material_bulk_file = open('Data/material_bulk_database.json', 'r')
        self.material_bulk_data = json.loads(material_bulk_file.read())
        material_bulk_file.close()

        self.NMR_weightsFile = 'Data\\NMR_weights.xlsx'

        self.nmrDataTools = NMRData()
        
        self.CalibrationLine_AcomArea()
        self.ComperisonPlots()
        self.SurfaceAreaCalculation()
        self.CalibrationLine_Spinsolve()





    #************************************************************************#
    ######################    TAB 1    #######################################
    #************************************************************************#


    def CalibrationLine_AcomArea(self):
        
        #GUI Variables
        self.ui.dateTimeEdit_dateTime.setDateTime(QtCore.QDateTime.currentDateTime())
        self.ui.dateTimeEdit_dateTime.setDisplayFormat("dd/MM/yyyy")

        self.user = self.ui.plainTextEdit_user
        self.bulkName = self.ui.comboBox_bulkName
        self.dateTime = self.ui.dateTimeEdit_dateTime
        self.densityBulk = self.ui.plainTextEdit_densityBulk
        self.evaluation_double = self.ui.checkBox_evaluation_other
        self.evaluation_single = self.ui.checkBox_evaluation_single
        self.language_english = self.ui.checkBox_language_english
        self.language_german = self.ui.checkBox_language_german
        self.materialName = self.ui.comboBox_materialName
        self.particleDensity = self.ui.plainTextEdit_particleDensity
        self.remarks = self.ui.plainTextEdit_remarks
        self.surfaceArea_Argon = self.ui.plainTextEdit_surfaceArea_Argon
        self.temperature = self.ui.plainTextEdit_temperature



        self.user.setPlainText("Alexander Michalowski")
        #bulkName.addItems(["Milipore-Wasser_LFG", "Milipore-Wasser"])
        #densityBulk.setPlainText("1")
        #materialName.addItems(["80nmIV", "Test2"])
        #particleDensity.setPlainText("2.14")
        self.remarks.setPlainText("#76,#77,#79")
        #surfaceArea_Argon.setPlainText("47.0")
        self.temperature.setPlainText("25°C")

        self.MaterialDataLoad("start")
        self.BulkDataLoad("start")

        self.run_btn = self.ui.pushButton_run_MeasurementFile
        self.run_btn.clicked.connect(self.fetch_input)


        #ADD - DELETE FILES BUTTONS
        self.add_btn = self.ui.pushButton_AddMesurementFiles
        self.add_btn.clicked.connect(self.addFiles)

        self.remove_btn = self.ui.pushButton_RemoveAllMesurementFiles
        self.remove_btn.clicked.connect(self.removeFiles)
        
        self.removeSel_btn = self.ui.pushButton_RemoveSelectedMeasrementFiles
        self.removeSel_btn.clicked.connect(self.removeSelFiles)

        
        
        #COMBO BOXES
        self.comboMaterial = self.ui.comboBox_materialName
        self.comboMaterial.currentTextChanged.connect(self.setMaterialProperties)

        self.comboBulk = self.ui.comboBox_bulkName
        self.comboBulk.currentTextChanged.connect(self.setBulkProperties)

        self.comboWeights = self.ui.comboBox_selectWeighInData
        self.comboWeights_surfaceAreaCalculation = self.ui.comboBox_surfaceAreaCalculation_selectWeighInData
        self.comboWeights_Spinsolve = self.ui.comboBox_selectWeighInData_Spinsolve
        self.SheetComboLoad()

        
        #ADD - DELETE MAT/BULK BUTTONS
        
        self.addMaterial_btn = self.ui.pushButton_AddMaterial
        self.addMaterial_btn.clicked.connect(lambda:self.addMaterialData(self.ui.comboBox_materialName.currentText(), self.ui.plainTextEdit_surfaceArea_Argon.toPlainText(), self.ui.plainTextEdit_particleDensity.toPlainText()))

        self.addBulk_btn = self.ui.pushButton_AddBulk
        self.addBulk_btn.clicked.connect(lambda:self.addBulkData(self.ui.comboBox_bulkName.currentText(), self.ui.plainTextEdit_densityBulk.toPlainText()))

        self.removeMaterial_btn = self.ui.pushButton_RemoveMaterial
        self.removeMaterial_btn.clicked.connect(lambda:self.removeMaterialData(self.ui.comboBox_materialName.currentText()))

        self.removeBulk_btn = self.ui.pushButton_RemoveBulk
        self.removeBulk_btn.clicked.connect(lambda:self.removeBulkData(self.ui.comboBox_bulkName.currentText()))

        self.tableModel = TableModelData()

        self.updateWeights_btn = self.ui.pushButton_SetUpdateWeights
        self.updateWeights_btn.clicked.connect(self.updateWeights)

        self.updateSelWeights_btn = self.ui.pushButton_SetUpdateSelectedWeights
        self.updateSelWeights_btn.clicked.connect(self.updateSelWeights)



    def addMaterialData(self, materialName, surfaceAreaArgon, particleDensity):

        for i in self.material_bulk_data['materials']:
            if i['materialName'] == materialName:
                i['surfaceAreaArgon'] = surfaceAreaArgon
                i['particleDensity'] = particleDensity

                return 0
        
        new_entry = [{
            "materialName": materialName,
            "surfaceAreaArgon": surfaceAreaArgon,
            "particleDensity": particleDensity
        }]
        materiallist = self.material_bulk_data['materials']
        self.material_bulk_data['materials'].extend(new_entry)
        
        self.MaterialDataLoad("add")
        self.updateJsonDatabase()



    def addBulkData(self, bulkName, densityBulk):
        
        for i in self.material_bulk_data['bulk']:
            if i['bulkName'] == bulkName:
                i['densityBulk'] = densityBulk

                return 0
        
        new_entry = [{
            "bulkName": bulkName,
            "densityBulk": densityBulk,
        }]
        bulklist = self.material_bulk_data['bulk']
        self.material_bulk_data['bulk'].extend(new_entry)
        
        self.BulkDataLoad("add")
        self.updateJsonDatabase()

    def removeMaterialData(self, materialName):
        
        j=0
        for i in self.material_bulk_data['materials']:
            if i['materialName'] == materialName:
                self.material_bulk_data['materials'].pop(j)
            j += 1
        
        self.MaterialDataLoad("delete")
        self.updateJsonDatabase()


    def removeBulkData(self, bulkName):
        
        j=0
        for i in self.material_bulk_data['bulk']:
            if i['bulkName'] == bulkName:
                self.material_bulk_data['bulk'].pop(j)
            j += 1
        
        self.BulkDataLoad("delete")
        self.updateJsonDatabase()
    

    def updateJsonDatabase(self):
        
        jsonData = json.dumps(self.material_bulk_data)

        material_bulk_file = open('Data/material_bulk_database.json', 'w')
        material_bulk_file.write(jsonData)
        material_bulk_file.close()

    def MaterialDataLoad(self,state):

        material = []

        for i in self.material_bulk_data['materials']:

            material.append(i['materialName'])

        
        self.updateMaterialComboBoxes(material,state)
        

        self.setMaterialProperties()
        

    def BulkDataLoad(self,state):

        bulk = []

        for i in self.material_bulk_data['bulk']:

            bulk.append(i['bulkName'])

        self.updateBulkComboBoxes(bulk,state)
        self.setBulkProperties()

    def updateMaterialComboBoxes(self,material,state):
        
        matindex = 0

        if state == "add":
            matindex = self.materialName.count()

        self.materialName.clear()

        self.materialName.addItems(material)
        self.materialName.setCurrentIndex(matindex)
        self.setMaterialProperties()


        #Tab3
        materialName_surfaceAreaCalculation = self.ui.comboBox_surfaceAreaCalculation_materialName
        materialName_surfaceAreaCalculation.addItems(material)
        materialName_surfaceAreaCalculation.setCurrentIndex(matindex)
        self.setMaterialProperties_surfaceAreaCalculation()

        #Tab4
        materialName_Spinsolve = self.ui.comboBox_materialName_Spinsolve
        materialName_Spinsolve.addItems(material)
        materialName_Spinsolve.setCurrentIndex(matindex)
        self.setMaterialProperties_Spinsolve()


    def updateBulkComboBoxes(self, bulk, state):

        bulkindex = 0

        if state == "add":
            bulkindex = self.bulkName.count()

        self.bulkName.clear()

        self.bulkName.addItems(bulk)
        self.bulkName.setCurrentIndex(bulkindex)
        self.setBulkProperties()

        #Tab3
        bulkName_surfaceAreaCalculation = self.ui.comboBox_surfaceAreaCalculation_bulkName
        bulkName_surfaceAreaCalculation.addItems(bulk)
        bulkName_surfaceAreaCalculation.setCurrentIndex(bulkindex)
        self.setBulkProperties_surfaceAreaCalculation()

        #Tab4
        bulkName_Spinsolve = self.ui.comboBox_bulkName_Spinsolve
        bulkName_Spinsolve.addItems(bulk)
        bulkName_Spinsolve.setCurrentIndex(bulkindex)
        self.setBulkProperties_Spinsolve()


    def setMaterialProperties(self):
        
        matName = self.materialName.currentText()

        for i in self.material_bulk_data['materials']:

            if i['materialName'] == (matName):
                self.particleDensity.setPlainText(i['particleDensity'])
                self.surfaceArea_Argon.setPlainText(i['surfaceAreaArgon'])

    def setBulkProperties(self):
        
        bName = self.bulkName.currentText()

        for i in self.material_bulk_data['bulk']:

            if i['bulkName'] == (bName):
                self.densityBulk.setPlainText(i['densityBulk'])


    
    def SheetComboLoad(self):

        wb = load_workbook(self.NMR_weightsFile, read_only=True, keep_links=False)
        sheet_names = []
        for sheet_name in wb.sheetnames:
            sheet_names.append(sheet_name)
        self.comboWeights.addItems(sheet_names)
        self.comboWeights_surfaceAreaCalculation.addItems(sheet_names)   
        self.comboWeights_Spinsolve.addItems(sheet_names)


    def updateWeights(self):
        self.tableModel.updateAllWeights(self.comboWeights.currentText())
    def updateSelWeights(self):

        #fetch selected
        plainText_selectedWeights = self.ui.plainTextEdit_SetUpdateWeights
        selectedRowsUPD = plainText_selectedWeights.toPlainText().split(",")

        selectedRowsUPD = [int(i) for i in selectedRowsUPD]

        self.tableModel.updateSelWeights(self.comboWeights.currentText(), selectedRowsUPD)
        
    def removeSelFiles(self):

        #fetch selected
        plainText_selectedFiles = self.ui.plainTextEdit_RemoveSelectedFilesMeasrementFiles
        selectedRowsDEL = plainText_selectedFiles.toPlainText().split(",")

        selectedRowsDEL = [int(i) for i in selectedRowsDEL]

        self.tableModel.DelRows(selectedRowsDEL)
        
    #Grouping Files to cook them up for the script
    def groupFiles(self, files):

        fileInfo = list()
        #groupedfiles = []
        
        for f in files:
            results = self.nmrDataTools.getNMRinfo(f)
            results.append(f)
            fileInfo.append(results)

        #print(fileInfo)
        #Update UI
        groupedT1 = list()
        groupedT2 = list()

        #create groupedfiles for script input
        fileInfo.sort(key=lambda x: x[1])
        groupedBySampleName = functools.reduce(lambda l, x: (l.append([x]) if l[-1][0][1] != x[1] else l[-1].append(x)) or l, fileInfo[1:], [[fileInfo[0]]]) if fileInfo else []
        
        numberOfConcentrations = 0

        

        for groupRow in groupedBySampleName:

            numberOfConcentrations += 1
            #print("GROUP")    
            #print(groupRow)

            if len(groupRow) != 6:
                msgBox = QtWidgets.QMessageBox()
                msgBox.setIcon(QtWidgets.QMessageBox.Warning)
                msgBox.setText("Wrong Input!!")
                msgBox.setWindowTitle("Eisai malakas")
                msgBox.setStandardButtons(QtWidgets.QMessageBox.Ok)
                returnValue = msgBox.exec()
                return 0
            

            for groupT in groupRow:

                if str(groupT[0]) == "T2A":
                    groupedT2.append(groupT)
                elif str(groupT[0]) == "T1":
                    groupedT1.append(groupT)
                else:
                    return 0
        
        #print(groupedT1)
        #print(groupedT2)

        #self.CreateTable(self.numberOfConcentrations, self.groupedT1, self.groupedT2)

        sheetname = self.comboWeights.currentText()

        model = self.tableModel.AddRows(numberOfConcentrations, groupedT1, groupedT2, sheetname)
        
        table = self.ui.tableView_mesurementFiles
        table.setModel(model)
        table.horizontalHeader().resizeSection(0, 330)
        table.resizeColumnsToContents()    



    #ADD REMOVE FILES LOGIC
    def addFiles(self):
        textForOpeningFiles = "Please select Files for Concentration"
        files = fileopenbox(textForOpeningFiles, "Dunno", default = "m:/AcornArea/", filetypes= "*.txt", multiple=True)
        if files:
            #print(files)
            self.groupFiles(files)
        else:
            #print("exit")
            pass
        
    def removeFiles(self):
        self.tableModel.RemoveAllRows()
        

    


    #RUN BUTTON
    def fetch_input(self):
        
        user = self.ui.plainTextEdit_user
        bulkName = self.ui.comboBox_bulkName
        dateTime = self.ui.dateTimeEdit_dateTime
        densityBulk = self.ui.plainTextEdit_densityBulk
        evaluation_double = self.ui.checkBox_evaluation_other
        evaluation_single = self.ui.checkBox_evaluation_single
        language_english = self.ui.checkBox_language_english
        language_german = self.ui.checkBox_language_german
        materialName = self.ui.comboBox_materialName
        particleDensity = self.ui.plainTextEdit_particleDensity
        remarks = self.ui.plainTextEdit_remarks
        surfaceArea_Argon = self.ui.plainTextEdit_surfaceArea_Argon
        temperature = self.ui.plainTextEdit_temperature

        #print(dateTime.dateTime().toString(self.ui.dateTimeEdit_dateTime.displayFormat()))

        model = self.ui.tableView_mesurementFiles.model()

        groupedT1 = self.tableModel.getGroupedT1()
        groupedT2 = self.tableModel.getGroupedT2()
        numOfConcentrations = self.tableModel.getNumOfConcentrations()

        # print(groupedT1)
        # print(groupedT2)
        # print(type(numOfConcentrations))
        # print(model.rowCount())

        data = [[0 for x in range(model.columnCount())] for y in range(model.rowCount())]
        
        files_T1 = [[0 for x in range(3)] for y in range(model.rowCount())]
        files_T2 = [[0 for x in range(3)] for y in range(model.rowCount())]
        filespath_T1 = [[0 for x in range(3)] for y in range(model.rowCount())]
        filespath_T2 = [[0 for x in range(3)] for y in range(model.rowCount())]
        
        liquidmassfromTable = list()
        particlemassfromTable = list()
        #print(particlemassfromTable)
        group_row = 0

        for row in range(model.rowCount()):
            #data.append([])

            pos = int(model.data( model.index(row, 0)))
            print(pos)
            if pos < 1 or pos > model.rowCount():
                msgBox = QtWidgets.QMessageBox()
                msgBox.setIcon(QtWidgets.QMessageBox.Warning)
                msgBox.setText("Wrong position given")
                msgBox.setWindowTitle("Eisai malakas")
                msgBox.setStandardButtons(QtWidgets.QMessageBox.Ok)
                returnValue = msgBox.exec()
                return 0

            #Populate filenames in the correct order
            
            for i in range(3):
                files_T1[pos-1][i] = groupedT1[group_row][2]
                files_T2[pos-1][i] = groupedT2[group_row][2]
                filespath_T1[pos-1][i] = groupedT1[group_row][4]
                filespath_T2[pos-1][i] = groupedT2[group_row][4]
                group_row += 1

            for column in range(model.columnCount()):
                #print("ooooo "+str(column))
                index = model.index(row, column)
                # We suppose data are strings
                data[pos-1][column] = str(model.data(index)) 

        for j in range(model.rowCount()) :
            liquidmassfromTable.append(float(data[j][4]))
            particlemassfromTable.append(float(data[j][5]))


        allFiles = files_T1 + files_T2

        #print(liquidmassfromTable)
        #print(particlemassfromTable)
        print(allFiles)
        print(filespath_T1)
        print(filespath_T2)
        #print(data)
        
        if evaluation_single.isChecked() == True and evaluation_double.isChecked() == False:
            evaluation = "single"
        elif evaluation_single.isChecked() == False and evaluation_double.isChecked() == True:
            evaluation = "????"
        else:
            msgBox = QtWidgets.QMessageBox()
            msgBox.setIcon(QtWidgets.QMessageBox.Warning)
            msgBox.setText("both checked evaluation")
            msgBox.setWindowTitle("Eisai malakas")
            msgBox.setStandardButtons(QtWidgets.QMessageBox.Ok)
            returnValue = msgBox.exec()
            return 0
        
        if language_english.isChecked() == True and language_german.isChecked() == False:
            language = "english"
        elif language_english.isChecked() == False and language_german.isChecked() == True:
            language = "german"
        else:
            language = "english"
            driver = Driver_referenceMeasurement_createFile_importExcel()
            driver.runDriver(materialName.currentText(), evaluation, bulkName.currentText(), user.toPlainText(), language, remarks.toPlainText(), temperature.toPlainText(), float(surfaceArea_Argon.toPlainText()), float(densityBulk.toPlainText()), float(particleDensity.toPlainText()), dateTime.dateTime().toString("yyyyMMdd"), numOfConcentrations, files_T1, files_T2, filespath_T1, filespath_T2, liquidmassfromTable, particlemassfromTable )
            language = "german"

        driver = Driver_referenceMeasurement_createFile_importExcel()
        driver.runDriver(materialName.currentText(), evaluation, bulkName.currentText(), user.toPlainText(), language, remarks.toPlainText(), temperature.toPlainText(), float(surfaceArea_Argon.toPlainText()), float(densityBulk.toPlainText()), float(particleDensity.toPlainText()), dateTime.dateTime().toString("yyyyMMdd"), numOfConcentrations, files_T1, files_T2, filespath_T1, filespath_T2, liquidmassfromTable, particlemassfromTable )
        




    #************************************************************************#
    ######################    TAB 4    #######################################
    #************************************************************************#


    def CalibrationLine_Spinsolve(self):
            
        #GUI Variables
        self.ui.dateTimeEdit_dateTime_Spinsolve.setDateTime(QtCore.QDateTime.currentDateTime())
        self.ui.dateTimeEdit_dateTime_Spinsolve.setDisplayFormat("dd/MM/yyyy")

        self.user_Spinsolve = self.ui.plainTextEdit_user_Spinsolve
        
        self.dateTime_Spinsolve = self.ui.dateTimeEdit_dateTime_Spinsolve
        self.densityBulk_Spinsolve = self.ui.plainTextEdit_densityBulk_Spinsolve
        self.evaluation_double_Spinsolve = self.ui.checkBox_evaluation_other_Spinsolve
        self.evaluation_single_Spinsolve = self.ui.checkBox_evaluation_single_Spinsolve
        self.language_english_Spinsolve = self.ui.checkBox_language_english_Spinsolve
        self.language_german_Spinsolve = self.ui.checkBox_language_german_Spinsolve
        
        self.particleDensity_Spinsolve = self.ui.plainTextEdit_particleDensity_Spinsolve
        self.remarks_Spinsolve = self.ui.plainTextEdit_remarks_Spinsolve
        self.surfaceArea_Argon_Spinsolve = self.ui.plainTextEdit_surfaceArea_Argon_Spinsolve
        self.temperature_Spinsolve = self.ui.plainTextEdit_temperature_Spinsolve

        self.materialName_Spinsolve = self.ui.comboBox_materialName_Spinsolve
        self.bulkName_Spinsolve = self.ui.comboBox_bulkName_Spinsolve



        self.user_Spinsolve.setPlainText("Alexander Michalowski")
        #bulkName_Spinsolve.addItems(["Milipore-Wasser_LFG", "Milipore-Wasser"])
        #densityBulk_Spinsolve.setPlainText("1")
        #materialName_Spinsolve.addItems(["80nmIV", "Test2"])
        #particleDensity_Spinsolve.setPlainText("2.14")
        self.remarks_Spinsolve.setPlainText("#76,#77,#79")
        #surfaceArea_Argon_Spinsolve.setPlainText("47.0")
        self.temperature_Spinsolve.setPlainText("25°C")

        self.run_btn_Spinsolve = self.ui.pushButton_run_MeasurementFile_Spinsolve
        self.run_btn_Spinsolve.clicked.connect(self.fetch_input_Spinsolve)


        #ADD - DELETE FILES BUTTONS
        self.add_btn_Spinsolve = self.ui.pushButton_AddMesurementFiles_Spinsolve
        self.add_btn_Spinsolve.clicked.connect(self.addFiles_Spinsolve)

        self.remove_btn_Spinsolve = self.ui.pushButton_RemoveAllMesurementFiles_Spinsolve
        self.remove_btn_Spinsolve.clicked.connect(self.removeFiles_Spinsolve)
        
        self.removeSel_btn_Spinsolve = self.ui.pushButton_RemoveSelectedMeasrementFiles_Spinsolve
        self.removeSel_btn_Spinsolve.clicked.connect(self.removeSelFiles_Spinsolve)

        
        
        #COMBO BOXES
        self.comboMaterial_Spinsolve = self.ui.comboBox_materialName_Spinsolve
        self.comboMaterial_Spinsolve.currentTextChanged.connect(self.setMaterialProperties_Spinsolve)

        self.comboBulk_Spinsolve = self.ui.comboBox_bulkName_Spinsolve
        self.comboBulk_Spinsolve.currentTextChanged.connect(self.setBulkProperties_Spinsolve)

        self.comboWeights_Spinsolve = self.ui.comboBox_selectWeighInData_Spinsolve
        #self.comboWeights_surfaceAreaCalculation = self.ui.comboBox_surfaceAreaCalculation_selectWeighInData

        #ADD - DELETE MAT/BULK BUTTONS
        
        self.addMaterial_btn_Spinsolve = self.ui.pushButton_AddMaterial_Spinsolve
        self.addMaterial_btn_Spinsolve.clicked.connect(lambda:self.addMaterialData(self.ui.comboBox_materialName_Spinsolve.currentText(), self.ui.plainTextEdit_surfaceArea_Argon_Spinsolve.toPlainText(), self.ui.plainTextEdit_particleDensity_Spinsolve.toPlainText()))

        self.addBulk_btn_Spinsolve = self.ui.pushButton_AddBulk_Spinsolve
        self.addBulk_btn_Spinsolve.clicked.connect(lambda:self.addBulkData(self.ui.comboBox_bulkName_Spinsolve.currentText(), self.ui.plainTextEdit_densityBulk_Spinsolve.toPlainText()))

        self.removeMaterial_btn_Spinsolve = self.ui.pushButton_RemoveMaterial_Spinsolve
        self.removeMaterial_btn_Spinsolve.clicked.connect(lambda:self.removeMaterialData(self.ui.comboBox_materialName_Spinsolve.currentText()))

        self.removeBulk_btn_Spinsolve = self.ui.pushButton_RemoveBulk_Spinsolve
        self.removeBulk_btn_Spinsolve.clicked.connect(lambda:self.removeBulkData(self.ui.comboBox_bulkName_Spinsolve.currentText()))

        self.tableModel_Spinsolve = TableModelData()

        self.updateWeights_Spinsolve_btn = self.ui.pushButton_SetUpdateWeights_Spinsolve
        self.updateWeights_Spinsolve_btn.clicked.connect(self.updateWeights_Spinsolve)

        self.updateSelWeights_Spinsolve_btn = self.ui.pushButton_SetUpdateSelectedWeights_Spinsolve
        self.updateSelWeights_Spinsolve_btn.clicked.connect(self.updateSelWeights_Spinsolve)





    def setMaterialProperties_Spinsolve(self):
        
        matName = self.ui.comboBox_materialName_Spinsolve.currentText()

        for i in self.material_bulk_data['materials']:

            if i['materialName'] == (matName):
                self.ui.plainTextEdit_particleDensity_Spinsolve.setPlainText(i['particleDensity'])
                self.ui.plainTextEdit_surfaceArea_Argon_Spinsolve.setPlainText(i['surfaceAreaArgon'])

    def setBulkProperties_Spinsolve(self):
        
        bName = self.ui.comboBox_bulkName_Spinsolve.currentText()

        for i in self.material_bulk_data['bulk']:

            if i['bulkName'] == (bName):
                self.ui.plainTextEdit_densityBulk_Spinsolve.setPlainText(i['densityBulk'])




    def updateWeights_Spinsolve(self):
        self.tableModel_Spinsolve.updateAllWeights(self.comboWeights_Spinsolve.currentText())
    def updateSelWeights_Spinsolve(self):

        #fetch selected
        plainText_selectedWeights = self.ui.plainTextEdit_SetUpdateWeights_Spinsolve
        selectedRowsUPD = plainText_selectedWeights.toPlainText().split(",")

        selectedRowsUPD = [int(i) for i in selectedRowsUPD]

        self.tableModel_Spinsolve.updateSelWeights(self.comboWeights_Spinsolve.currentText(), selectedRowsUPD)
        
    def removeSelFiles_Spinsolve(self):

        #fetch selected
        plainText_selectedFiles = self.ui.plainTextEdit_RemoveSelectedFilesMeasrementFiles_Spinsolve
        selectedRowsDEL = plainText_selectedFiles.toPlainText().split(",")

        selectedRowsDEL = [int(i) for i in selectedRowsDEL]

        self.tableModel_Spinsolve.DelRows(selectedRowsDEL)


    #ADD REMOVE FILES LOGIC
    def addFiles_Spinsolve(self):

        T1_fileName = 'data_graph.csv'
        T2_fileName = 'spectrum_processed.csv'

        T1_files = list()
        T2_files = list()

        textForOpeningFiles = "Please select Folder for Concentration"
        #files = fileopenbox(textForOpeningFiles, "Dunno", default = "", filetypes= "*.txt", multiple=True)
        dir = diropenbox(textForOpeningFiles, "Dunno", default = "../Spinsolve")

        if dir:
            i = 0 
            j = 0
            for (path,dirs,files) in os.walk(dir):
                
                # print(path,'Path')
                # print(dirs,'Dir')
                # print(files,'Files')
                
                if T1_fileName in files:

                    T1_files.append([])

                    T1_files[i].append('T1')
                    T1_files[i].append(dir.split('\\')[-1])
                    T1_files[i].append(path.split('\\')[-1])

                    filepath = path+'\\'+T1_fileName
                    T1_files[i].append(self.nmrDataTools.SpinsolveT1_graphdata(filepath))
                    T1_files[i].append(filepath)

                    i += 1

                elif T2_fileName in files:

                    T2_files.append([])

                    T2_files[j].append('T2')
                    T2_files[j].append(dir.split('\\')[-1])
                    T2_files[j].append(path.split('\\')[-1])

                    filepath = path+'\\'+T2_fileName
                    T2_files[j].append(self.nmrDataTools.SpinsolveT2_log(filepath))
                    T2_files[j].append(filepath)

                    j += 1

            
            #print(T1_files,'T1')
            #print(T2_files,'T2')

            sheetname = self.comboWeights_Spinsolve.currentText()

            model = self.tableModel_Spinsolve.AddRows(1, T1_files, T2_files, sheetname)
            
            table = self.ui.tableView_mesurementFiles_Spinsolve
            table.setModel(model)
            table.horizontalHeader().resizeSection(0, 330)
            table.resizeColumnsToContents()    
        else:
            #print("exit")
            pass
        
    def removeFiles_Spinsolve(self):
        self.tableModel_Spinsolve.RemoveAllRows()
        




    #RUN BUTTON
    def fetch_input_Spinsolve(self):
        
        user_Spinsolve = self.ui.plainTextEdit_user_Spinsolve
        bulkName_Spinsolve = self.ui.comboBox_bulkName_Spinsolve
        dateTime = self.ui.dateTimeEdit_dateTime_Spinsolve
        densityBulk_Spinsolve = self.ui.plainTextEdit_densityBulk_Spinsolve
        evaluation_double_Spinsolve = self.ui.checkBox_evaluation_other
        evaluation_single_Spinsolve = self.ui.checkBox_evaluation_single_Spinsolve
        language_english_Spinsolve = self.ui.checkBox_language_english_Spinsolve
        language_german_Spinsolve = self.ui.checkBox_language_german_Spinsolve
        materialName_Spinsolve = self.ui.comboBox_materialName_Spinsolve
        particleDensity_Spinsolve = self.ui.plainTextEdit_particleDensity_Spinsolve
        remarks_Spinsolve = self.ui.plainTextEdit_remarks_Spinsolve
        surfaceArea_Argon_Spinsolve = self.ui.plainTextEdit_surfaceArea_Argon_Spinsolve
        temperature_Spinsolve = self.ui.plainTextEdit_temperature_Spinsolve

        #print(dateTime.dateTime().toString(self.ui.dateTimeEdit_dateTime_Spinsolve.displayFormat()))

        model = self.ui.tableView_mesurementFiles_Spinsolve.model()

        groupedT1 = self.tableModel_Spinsolve.getGroupedT1()
        groupedT2 = self.tableModel_Spinsolve.getGroupedT2()
        numOfConcentrations = self.tableModel_Spinsolve.getNumOfConcentrations()

        # print(groupedT1)
        # print(groupedT2)
        # print(type(numOfConcentrations))
        # print(model.rowCount())

        data = [[0 for x in range(model.columnCount())] for y in range(model.rowCount())]
        
        files_T1 = [[0 for x in range(3)] for y in range(model.rowCount())]
        files_T2 = [[0 for x in range(3)] for y in range(model.rowCount())]
        filespath_T1 = [[0 for x in range(3)] for y in range(model.rowCount())]
        filespath_T2 = [[0 for x in range(3)] for y in range(model.rowCount())]

        value_T1 = [[0 for x in range(3)] for y in range(model.rowCount())]
        value_T2 = [[0 for x in range(3)] for y in range(model.rowCount())]

        
        liquidmassfromTable = list()
        particlemassfromTable = list()
        #print(particlemassfromTable)
        group_row = 0

        for row in range(model.rowCount()):
            #data.append([])

            pos = int(model.data( model.index(row, 0)))
            #print(pos)
            if pos < 1 or pos > model.rowCount():
                msgBox = QtWidgets.QMessageBox()
                msgBox.setIcon(QtWidgets.QMessageBox.Warning)
                msgBox.setText("Wrong position given")
                msgBox.setWindowTitle("Eisai malakas")
                msgBox.setStandardButtons(QtWidgets.QMessageBox.Ok)
                returnValue = msgBox.exec()
                return 0

            #Populate filenames in the correct order
            
            for i in range(3):
                files_T1[pos-1][i] = groupedT1[group_row][2]
                files_T2[pos-1][i] = groupedT2[group_row][2]
                filespath_T1[pos-1][i] = groupedT1[group_row][4]
                filespath_T2[pos-1][i] = groupedT2[group_row][4]

                value_T1[pos-1][i] = groupedT1[group_row][3]
                value_T2[pos-1][i] = groupedT2[group_row][3]

                group_row += 1

            for column in range(model.columnCount()):
                #print("ooooo "+str(column))
                index = model.index(row, column)
                # We suppose data are strings
                data[pos-1][column] = str(model.data(index)) 

        for j in range(model.rowCount()) :
            liquidmassfromTable.append(float(data[j][4]))
            particlemassfromTable.append(float(data[j][5]))


        allFiles = files_T1 + files_T2

        # print(liquidmassfromTable)
        # print(particlemassfromTable)
        # print(allFiles)
        # print(filespath_T1)
        # print(filespath_T2)
        # print(value_T1)
        # print(value_T2)
        #print(data)
        
        if evaluation_single_Spinsolve.isChecked() == True and evaluation_double_Spinsolve.isChecked() == False:
            evaluation = "single"
        elif evaluation_single_Spinsolve.isChecked() == False and evaluation_double_Spinsolve.isChecked() == True:
            evaluation = "????"
        else:
            msgBox = QtWidgets.QMessageBox()
            msgBox.setIcon(QtWidgets.QMessageBox.Warning)
            msgBox.setText("both checked evaluation")
            msgBox.setWindowTitle("Eisai malakas")
            msgBox.setStandardButtons(QtWidgets.QMessageBox.Ok)
            returnValue = msgBox.exec()
            return 0
        
        if language_english_Spinsolve.isChecked() == True and language_german_Spinsolve.isChecked() == False:
            language = "english"
        elif language_english_Spinsolve.isChecked() == False and language_german_Spinsolve.isChecked() == True:
            language = "german"
        else:
            language = "english"
            driver = Driver_referenceMeasurement_createFile_withoutExcel()
            driver.runDriver(materialName_Spinsolve.currentText(), evaluation, bulkName_Spinsolve.currentText(), user_Spinsolve.toPlainText(), language, remarks_Spinsolve.toPlainText(), temperature_Spinsolve.toPlainText(), float(surfaceArea_Argon_Spinsolve.toPlainText()), float(densityBulk_Spinsolve.toPlainText()), float(particleDensity_Spinsolve.toPlainText()), dateTime.dateTime().toString("yyyyMMdd"), numOfConcentrations, value_T1, value_T2, filespath_T1, filespath_T2, liquidmassfromTable, particlemassfromTable )
            language = "german"

        driver = Driver_referenceMeasurement_createFile_withoutExcel()
        driver.runDriver(materialName_Spinsolve.currentText(), evaluation, bulkName_Spinsolve.currentText(), user_Spinsolve.toPlainText(), language, remarks_Spinsolve.toPlainText(), temperature_Spinsolve.toPlainText(), float(surfaceArea_Argon_Spinsolve.toPlainText()), float(densityBulk_Spinsolve.toPlainText()), float(particleDensity_Spinsolve.toPlainText()), dateTime.dateTime().toString("yyyyMMdd"), numOfConcentrations, value_T1, value_T2, filespath_T1, filespath_T2, liquidmassfromTable, particlemassfromTable )
        



    #************************************************************************#
    ######################    TAB 2    #######################################
    #************************************************************************#


    def ComperisonPlots(self):

        self.comperisonPlots_language_german = self.ui.checkBox_comperisonPlots_language_german
        self.comperisonPlots_language_english = self.ui.checkBox_comperisonPlots_language_english

        self.comperisonPlots_T1 = self.ui.checkBox_comperisonPlots_T1
        self.comperisonPlots_T2 = self.ui.checkBox_comperisonPlots_T2

        self.comperisonPlots_T1_PlotName = self.ui.plainTextEdit_comperisonPlots_T1_PlotName
        self.comperisonPlots_T2_PlotName = self.ui.plainTextEdit_comperisonPlots_T2_PlotName

        self.comperisonPlots_T1_PlotName.setPlainText('Sample T1')
        self.comperisonPlots_T2_PlotName.setPlainText('Sample T2')
        
        self.comperisonPlots_T1_calculationOfVolumeFraction = self.ui.plainTextEdit_comperisonPlots_T1_calculationOfVolumeFraction
        self.comperisonPlots_T2_calculationOfVolumeFraction = self.ui.plainTextEdit_comperisonPlots_T2_calculationOfVolumeFraction

        self.comperisonPlots_T1_calculationOfVolumeFraction.setPlainText('mass')
        self.comperisonPlots_T2_calculationOfVolumeFraction.setPlainText('mass')

        self.ui.pushButton_comperisonPlots_AddReferenceReferenceMesurementFiles
        self.ui.pushButton_comperisonPlots_RemoveAllReferenceMesurementFiles
        self.ui.pushButton_comperisonPlots_RemoveSelectedReferenceMeasrementFiles

        self.run_btn_comperisonPlots = self.ui.pushButton_ComerisonPlots_run
        self.run_btn_comperisonPlots.clicked.connect(self.fetch_input_comperisonPlots)


        #ADD - DELETE FILES BUTTONS
        self.add_btn_comperisonPlots = self.ui.pushButton_comperisonPlots_AddReferenceReferenceMesurementFiles
        self.add_btn_comperisonPlots.clicked.connect(self.addFiles_comperisonPlots)

        self.remove_btn_comperisonPlots = self.ui.pushButton_comperisonPlots_RemoveAllReferenceMesurementFiles
        self.remove_btn_comperisonPlots.clicked.connect(self.removeFiles_comperisonPlots)
        
        self.removeSel_btn_comperisonPlots = self.ui.pushButton_comperisonPlots_RemoveSelectedReferenceMeasrementFiles
        self.removeSel_btn_comperisonPlots.clicked.connect(self.removeSelFiles_comperisonPlots)


        self.comperisonPlots_RemoveSelectedFilesReferenceMeasrementFiles = self.ui.plainTextEdit_comperisonPlots_RemoveSelectedFilesReferenceMeasrementFiles

        self.tableView_comperisonPlots = self.ui.tableView_comperisonPlots_referenceMeasurementFiles

        self.tableModel_comperisonPlots = ComparisonTableModelData()

    def addFiles_comperisonPlots(self):

        textForOpeningFiles = "Please select Files for Comparison"
        files = fileopenbox(textForOpeningFiles, "Dunno", default = "../surfaceRelaxivity/", filetypes= "*.txt", multiple=True)
        if files:
            filepaths = []
            for file in files:
                filepaths.append(file.replace('\\','\\\\'))

            model = self.tableModel_comperisonPlots.AddRows(filepaths)

            self.tableView_comperisonPlots.setModel(model)
            self.tableView_comperisonPlots.horizontalHeader().resizeSection(0, 330)
            self.tableView_comperisonPlots.resizeColumnsToContents() 

        else:
            pass
    
    def removeFiles_comperisonPlots(self):
        self.tableModel_comperisonPlots.RemoveAllRows()

    def removeSelFiles_comperisonPlots(self):

        #fetch selected
        selectedRowsDEL =  self.comperisonPlots_RemoveSelectedFilesReferenceMeasrementFiles.toPlainText().split(",")

        selectedRowsDEL = [int(i) for i in selectedRowsDEL]
        

        self.tableModel_comperisonPlots.DelRows(selectedRowsDEL)
        
    # def removeReferenceMesurementFile(self):
    #     self.modelReference.removeRow(0)

    def fetch_input_comperisonPlots(self):
        
        model = self.tableView_comperisonPlots.model()

        data = [[0 for x in range(model.columnCount())] for y in range(model.rowCount())]
        
        files = list()
        legends = list()

        #fetch table
        for row in range(model.rowCount()):

            pos = int(model.data(model.index(row, 2)))

            if pos < 1 or pos > model.rowCount():
                msgBox = QtWidgets.QMessageBox()
                msgBox.setIcon(QtWidgets.QMessageBox.Warning)
                msgBox.setText("Wrong position given")
                msgBox.setWindowTitle("Eisai malakas")
                msgBox.setStandardButtons(QtWidgets.QMessageBox.Ok)
                returnValue = msgBox.exec()
                return 0

            for column in range(model.columnCount()):
                #print("ooooo "+str(column))
                index = model.index(row, column)
                # We suppose data are strings
                data[pos-1][column] = str(model.data(index))
        
        for j in range(model.rowCount()):
            legends.append(data[j][1])
        
        files = self.tableModel_comperisonPlots.getRelaxativityFiles()

        print(files)
        print(legends)


        PlotName = []
        PlotName.append(self.comperisonPlots_T1_PlotName.toPlainText())
        PlotName.append(self.comperisonPlots_T2_PlotName.toPlainText())

        VolumeFraction = []
        VolumeFraction.append(self.comperisonPlots_T1_calculationOfVolumeFraction.toPlainText())
        VolumeFraction.append(self.comperisonPlots_T2_calculationOfVolumeFraction.toPlainText())

        #T1 - T2 CheckBox
        if self.comperisonPlots_T1.isChecked() == True and self.comperisonPlots_T2.isChecked() == False:
            T1_T2 = "T1"
        elif self.comperisonPlots_T2.isChecked() == False and self.comperisonPlots_T1.isChecked() == True:
            T1_T2 = "T2"
        elif self.comperisonPlots_T2.isChecked() == True and self.comperisonPlots_T1.isChecked() == True:
            T1_T2 = "both"
        else:
            msgBox = QtWidgets.QMessageBox()
            msgBox.setIcon(QtWidgets.QMessageBox.Warning)
            msgBox.setText("Check evalutation T1, T2 or both")
            msgBox.setWindowTitle("Eisai malakas")
            msgBox.setStandardButtons(QtWidgets.QMessageBox.Ok)
            returnValue = msgBox.exec()
            return 0

        #Language CheckBox
        if self.comperisonPlots_language_english.isChecked() == True and self.comperisonPlots_language_german.isChecked() == False:
            language = "english"
        elif self.comperisonPlots_language_english.isChecked() == False and self.comperisonPlots_language_german.isChecked() == True:
            language = "german"
        else:
            language = "english"
            driver = Driver_comparisonPlots()
            driver.runDriver(files, T1_T2, VolumeFraction, legends, language, PlotName)
            language = "german"

        driver = Driver_comparisonPlots()
        driver.runDriver(files, T1_T2, VolumeFraction, legends, language, PlotName)

    #************************************************************************#
    ######################    TAB 3    #######################################
    #************************************************************************#


    def SurfaceAreaCalculation(self):
        
        #GUI Variables
        self.ui.dateTimeEdit_surfaceAreaCalculation_dateTime.setDateTime(QtCore.QDateTime.currentDateTime())
        self.ui.dateTimeEdit_surfaceAreaCalculation_dateTime.setDisplayFormat("dd/MM/yyyy")

        self.surfaceAreaCalculation_user = self.ui.plainTextEdit_surfaceAreaCalculation_user
        self.surfaceAreaCalculation_bulkName = self.ui.comboBox_surfaceAreaCalculation_bulkName
        self.surfaceAreaCalculation_dateTime = self.ui.dateTimeEdit_surfaceAreaCalculation_dateTime
        self.surfaceAreaCalculation_densityBulk = self.ui.plainTextEdit_surfaceAreaCalculation_densityBulk
        self.OneRelaxationTime = self.ui.checkBox_surfaceAreaCalculation_OneRelaxationTime
        self.TwoRelaxationTime = self.ui.checkBox_surfaceAreaCalculation_TwoRelaxationTime
        self.surfaceAreaCalculation_language_english = self.ui.checkBox_surfaceAreaCalculation_language_english
        self.surfaceAreaCalculation_language_german = self.ui.checkBox_surfaceAreaCalculation_language_german
        self.surfaceAreaCalculation_materialName = self.ui.comboBox_surfaceAreaCalculation_materialName
        self.surfaceAreaCalculation_particleDensity = self.ui.plainTextEdit_surfaceAreaCalculation_particleDensity
        self.surfaceAreaCalculation_remarks = self.ui.plainTextEdit_surfaceAreaCalculation_remarks
        self.surfaceAreaCalculation_surfaceArea_Argon = self.ui.plainTextEdit_surfaceAreaCalculation_surfaceArea_Argon
        self.surfaceAreaCalculation_temperature = self.ui.plainTextEdit_surfaceAreaCalculation_temperature

        self.surfaceAreaCalculation_user.setPlainText("Alexander Michalowski")
        #surfaceAreaCalculation_bulkName.addItems(["Milipore-Wasser_LFG", "Milipore-Wasser"])
        #surfaceAreaCalculation_densityBulk.setPlainText("1")
        #surfaceAreaCalculation_materialName.addItems(["80nmIV", "Test2"])
        #surfaceAreaCalculation_particleDensity.setPlainText("2.14")
        self.surfaceAreaCalculation_remarks.setPlainText("#76,#77,#79")
        #surfaceAreaCalculation_surfaceArea_Argon.setPlainText("47.0")
        self.surfaceAreaCalculation_temperature.setPlainText("25°C")


        self.run_btn_surfaceAreaCalculation = self.ui.pushButton_surfaceAreaCalculation_run
        self.run_btn_surfaceAreaCalculation.clicked.connect(self.fetch_input_surfaceAreaCalculation)


        #ADD - DELETE FILES BUTTONS
        self.add_btn_surfaceAreaCalculation = self.ui.pushButton_surfaceAreaCalculation_AddMesurementFiles
        self.add_btn_surfaceAreaCalculation.clicked.connect(self.addFiles_surfaceAreaCalculation)

        self.remove_btn_surfaceAreaCalculation = self.ui.pushButton_surfaceAreaCalculation_RemoveAllMesurementFiles
        self.remove_btn_surfaceAreaCalculation.clicked.connect(self.removeFiles_surfaceAreaCalculation)
        
        self.removeSel_btn_surfaceAreaCalculation = self.ui.pushButton_surfaceAreaCalculation_RemoveSelectedMeasrementFiles
        self.removeSel_btn_surfaceAreaCalculation.clicked.connect(self.removeSelFiles_surfaceAreaCalculation)

        
        
        #COMBO BOXES
        self.comboMaterial_surfaceAreaCalculation = self.ui.comboBox_surfaceAreaCalculation_materialName
        self.comboMaterial_surfaceAreaCalculation.currentTextChanged.connect(self.setMaterialProperties_surfaceAreaCalculation)

        self.comboBulk_surfaceAreaCalculation = self.ui.comboBox_surfaceAreaCalculation_bulkName
        self.comboBulk_surfaceAreaCalculation.currentTextChanged.connect(self.setBulkProperties_surfaceAreaCalculation)

        #ADD - DELETE MAT/BULK BUTTONS
        
        self.addMaterial_btn_surfaceAreaCalculation = self.ui.pushButton_surfaceAreaCalculation_AddMaterial
        self.addMaterial_btn_surfaceAreaCalculation.clicked.connect(lambda:self.addMaterialData(self.ui.comboBox_surfaceAreaCalculation_materialName.currentText(), self.ui.plainTextEdit_surfaceAreaCalculation_surfaceArea_Argon.toPlainText(), self.ui.plainTextEdit_surfaceAreaCalculation_particleDensity.toPlainText()))

        self.addBulk_btn_surfaceAreaCalculation = self.ui.pushButton_surfaceAreaCalculation_AddBulk
        self.addBulk_btn_surfaceAreaCalculation.clicked.connect(lambda:self.addBulkData(self.ui.comboBox_surfaceAreaCalculation_bulkName.currentText(), self.ui.plainTextEdit_surfaceAreaCalculation_densityBulk.toPlainText()))

        self.removeMaterial_btn_surfaceAreaCalculation = self.ui.pushButton_surfaceAreaCalculation_RemoveMaterial
        self.removeMaterial_btn_surfaceAreaCalculation.clicked.connect(lambda:self.removeMaterialData(self.ui.comboBox_surfaceAreaCalculation_materialName.currentText()))

        self.removeBulk_btn_surfaceAreaCalculation = self.ui.pushButton_surfaceAreaCalculation_RemoveBulk
        self.removeBulk_btn_surfaceAreaCalculation.clicked.connect(lambda:self.removeBulkData(self.ui.comboBox_surfaceAreaCalculation_bulkName.currentText()))

        self.tableModel_surfaceAreaCalculation = TableModelData()

        self.addReferenceMesurementFile_btn = self.ui.pushButton_surfaceAreaCalculation_AddReferenceMesurementFile
        self.addReferenceMesurementFile_btn.clicked.connect(self.addReferenceMeasrementFiles)

        self.removeReferenceMeasrementFile_btn = self.ui.pushButton_surfaceAreaCalculation_RemoveReferenceMeasrementFile
        self.removeReferenceMeasrementFile_btn.clicked.connect(self.removeReferenceMesurementFile)
        
        self.tableReferenceMesurementFiles = self.ui.tableView_surfaceAreaCalculation_ReferenceMesurementFiles

        self.updateWeights_btn_surfaceAreaCalculation = self.ui.pushButton_surfaceAreaCalculation_SetUpdateWeights
        self.updateWeights_btn_surfaceAreaCalculation.clicked.connect(self.updateWeights_surfaceAreaCalculation)

        self.updateSelWeights_btn_surfaceAreaCalculation = self.ui.pushButton_surfaceAreaCalculation_SetUpdateSelectedWeights
        self.updateSelWeights_btn_surfaceAreaCalculation.clicked.connect(self.updateSelWeights_surfaceAreaCalculation)

        

    def addReferenceMeasrementFiles(self):
        textForOpeningFiles = "Please select Files for Concentration"
        files = fileopenbox(textForOpeningFiles, "Dunno", default = "../surfaceRelaxivity/", filetypes= "*.txt", multiple=True)
        if files:
            #print(files)

            self.modelReference = QtGui.QStandardItemModel(0,0)
            #model.setHorizontalHeaderLabels('Filename')
            self.modelReference.insertRow(self.modelReference.rowCount())
            
            self.referenceFilepath = files[0].replace('\\','\\\\')

            item = QtGui.QStandardItem(self.referenceFilepath)
            self.modelReference.setItem(0, 0, item)

            self.tableReferenceMesurementFiles.setModel(self.modelReference)
            self.tableReferenceMesurementFiles.horizontalHeader().hide()
            self.tableReferenceMesurementFiles.verticalHeader().hide()
            self.tableReferenceMesurementFiles.resizeColumnsToContents()

            data = pd.read_excel(self.referenceFilepath, sheet_name='general',header=None)

            self.materialName_ReferenceMesurementFiles = data[1][0]
            self.date_ReferenceMesurementFiles = data[1][5]


        else:
            pass
        
    def removeReferenceMesurementFile(self):
        self.modelReference.removeRow(0)

    def setMaterialProperties_surfaceAreaCalculation(self):
        
        matName = self.ui.comboBox_surfaceAreaCalculation_materialName.currentText()

        for i in self.material_bulk_data['materials']:

            if i['materialName'] == (matName):
                self.ui.plainTextEdit_surfaceAreaCalculation_particleDensity.setPlainText(i['particleDensity'])
                self.ui.plainTextEdit_surfaceAreaCalculation_surfaceArea_Argon.setPlainText(i['surfaceAreaArgon'])


    def setBulkProperties_surfaceAreaCalculation(self):
        
        bName = self.ui.comboBox_surfaceAreaCalculation_bulkName.currentText()

        for i in self.material_bulk_data['bulk']:

            if i['bulkName'] == (bName):
                self.ui.plainTextEdit_surfaceAreaCalculation_densityBulk.setPlainText(i['densityBulk'])

        
    def removeSelFiles_surfaceAreaCalculation(self):

        #fetch selected
        plainText_selectedFiles = self.ui.plainTextEdit_surfaceAreaCalculation_RemoveSelectedFilesMeasrementFiles
        selectedRowsDEL = plainText_selectedFiles.toPlainText().split(",")

        selectedRowsDEL = [int(i) for i in selectedRowsDEL]
        

        self.tableModel_surfaceAreaCalculation.DelRows(selectedRowsDEL)
        
    def updateWeights_surfaceAreaCalculation(self):
        self.tableModel_surfaceAreaCalculation.updateAllWeights(self.comboWeights_surfaceAreaCalculation.currentText())
    def updateSelWeights_surfaceAreaCalculation(self):

        #fetch selected
        plainText_selectedWeights = self.ui.plainTextEdit_surfaceAreaCalculation_SetUpdateWeights
        selectedRowsUPD = plainText_selectedWeights.toPlainText().split(",")

        selectedRowsUPD = [int(i) for i in selectedRowsUPD]

        self.tableModel_surfaceAreaCalculation.updateSelWeights(self.comboWeights_surfaceAreaCalculation.currentText(), selectedRowsUPD)
    
    #Grouping Files to cook them up for the script
    def groupFiles_surfaceAreaCalculation(self, files):

        fileInfo = list()
        #groupedfiles = []
        
        for f in files:
            results = self.nmrDataTools.getNMRinfo(f)
            results.append(f)
            fileInfo.append(results)

        #print(fileInfo)
        #Update UI

        groupedT1_surfaceAreaCalculation = list()
        groupedT2_surfaceAreaCalculation = list()

        #create groupedfiles for script input
        fileInfo.sort(key=lambda x: x[1])
        groupedBySampleName = functools.reduce(lambda l, x: (l.append([x]) if l[-1][0][1] != x[1] else l[-1].append(x)) or l, fileInfo[1:], [[fileInfo[0]]]) if fileInfo else []
        
        numberOfConcentrations_surfaceAreaCalculation = 0

        

        for groupRow in groupedBySampleName:

            numberOfConcentrations_surfaceAreaCalculation += 1
            #print("GROUP")    
            #print(groupRow)

            if len(groupRow) != 6:
                msgBox = QtWidgets.QMessageBox()
                msgBox.setIcon(QtWidgets.QMessageBox.Warning)
                msgBox.setText("Wrong Input!!")
                msgBox.setWindowTitle("Eisai malakas")
                msgBox.setStandardButtons(QtWidgets.QMessageBox.Ok)
                returnValue = msgBox.exec()
                return 0
            

            for groupT in groupRow:

                if str(groupT[0]) == "T2A":
                    groupedT2_surfaceAreaCalculation.append(groupT)
                elif str(groupT[0]) == "T1":
                    groupedT1_surfaceAreaCalculation.append(groupT)
                else:
                    return 0

        #self.CreateTable(self.numberOfConcentrations, self.groupedT1, self.groupedT2)
        sheetname = self.comboWeights_surfaceAreaCalculation.currentText()

        model = self.tableModel_surfaceAreaCalculation.AddRows(numberOfConcentrations_surfaceAreaCalculation, groupedT1_surfaceAreaCalculation, groupedT2_surfaceAreaCalculation,sheetname)
        
        table = self.ui.tableView_surfaceAreaCalculation_mesurementFiles
        table.setModel(model)
        table.horizontalHeader().resizeSection(0, 330)
        table.resizeColumnsToContents()    





    #ADD REMOVE FILES LOGIC
    def addFiles_surfaceAreaCalculation(self):
        textForOpeningFiles = "Please select Files for Concentration"
        files = fileopenbox(textForOpeningFiles, "Dunno", default = "m:/AcornArea/", filetypes= "*.txt", multiple=True)
        if files:
            #print(files)
            self.groupFiles_surfaceAreaCalculation(files)
            


        else:
            #print("exit")
            pass
        
    def removeFiles_surfaceAreaCalculation(self):
        self.tableModel_surfaceAreaCalculation.RemoveAllRows()




    #RUN BUTTON
    def fetch_input_surfaceAreaCalculation(self):
        
        surfaceAreaCalculation_user = self.ui.plainTextEdit_surfaceAreaCalculation_user
        surfaceAreaCalculation_bulkName = self.ui.comboBox_surfaceAreaCalculation_bulkName
        surfaceAreaCalculation_dateTime = self.ui.dateTimeEdit_surfaceAreaCalculation_dateTime
        surfaceAreaCalculation_densityBulk = self.ui.plainTextEdit_surfaceAreaCalculation_densityBulk
        OneRelaxationTime = self.ui.checkBox_surfaceAreaCalculation_OneRelaxationTime
        TwoRelaxationTime = self.ui.checkBox_surfaceAreaCalculation_TwoRelaxationTime
        surfaceAreaCalculation_materialName = self.ui.comboBox_surfaceAreaCalculation_materialName
        surfaceAreaCalculation_particleDensity = self.ui.plainTextEdit_surfaceAreaCalculation_particleDensity
        surfaceAreaCalculation_remarks = self.ui.plainTextEdit_surfaceAreaCalculation_remarks
        surfaceAreaCalculation_surfaceArea_Argon = self.ui.plainTextEdit_surfaceAreaCalculation_surfaceArea_Argon
        surfaceAreaCalculation_temperature = self.ui.plainTextEdit_surfaceAreaCalculation_temperature

        #print(dateTime.dateTime().toString(self.ui.dateTimeEdit_surfaceAreaCalculation_dateTime.displayFormat()))

        model = self.ui.tableView_surfaceAreaCalculation_mesurementFiles.model()

        groupedT1 = self.tableModel_surfaceAreaCalculation.getGroupedT1()
        groupedT2 = self.tableModel_surfaceAreaCalculation.getGroupedT2()
        numOfConcentrations = self.tableModel_surfaceAreaCalculation.getNumOfConcentrations()

        #print("aaaa "+str(self.numberOfConcentrations_surfaceAreaCalculation))
        data = [[0 for x in range(model.columnCount())] for y in range(numOfConcentrations)]
        
        files_T1 = [[0 for x in range(3)] for y in range(numOfConcentrations)]
        files_T2 = [[0 for x in range(3)] for y in range(numOfConcentrations)]
        filespath_T1 = [[0 for x in range(3)] for y in range(numOfConcentrations)]
        filespath_T2 = [[0 for x in range(3)] for y in range(numOfConcentrations)]
        
        liquidmassfromTable = list()
        particlemassfromTable = list()
        group_row = 0

        for row in range(model.rowCount()):
            #data.append([])

            pos = int(model.data( model.index(row, 0)))
            #print(pos)
            if pos < 1 or pos > numOfConcentrations:
                msgBox = QtWidgets.QMessageBox()
                msgBox.setIcon(QtWidgets.QMessageBox.Warning)
                msgBox.setText("Wrong position given")
                msgBox.setWindowTitle("Eisai malakas")
                msgBox.setStandardButtons(QtWidgets.QMessageBox.Ok)
                returnValue = msgBox.exec()
                return 0

            #Populate filenames in the correct order
            
            for i in range(3):
                files_T1[pos-1][i] = groupedT1[group_row][2]
                files_T2[pos-1][i] = groupedT2[group_row][2]
                filespath_T1[pos-1][i] = groupedT1[group_row][4]
                filespath_T2[pos-1][i] = groupedT2[group_row][4]
                group_row += 1

            for column in range(model.columnCount()):
                index = model.index(row, column)
                # We suppose data are strings
                data[pos-1][column] = str(model.data(index)) 

        for j in range(numOfConcentrations) :
            liquidmassfromTable.append(float(data[j][4]))
            particlemassfromTable.append(float(data[j][5]))


        allFiles = files_T1 + files_T2

        # print(liquidmassfromTable)
        # print(particlemassfromTable)
        # print(allFiles)
        # print(data)
        
        if OneRelaxationTime.isChecked() == True and TwoRelaxationTime.isChecked() == False:
            Relaxation = "One"
        elif OneRelaxationTime.isChecked() == False and TwoRelaxationTime.isChecked() == True:
            Relaxation = "Two"
        else:
            msgBox = QtWidgets.QMessageBox()
            msgBox.setIcon(QtWidgets.QMessageBox.Warning)
            msgBox.setText("both checked Relaxation")
            msgBox.setWindowTitle("Eisai malakas")
            msgBox.setStandardButtons(QtWidgets.QMessageBox.Ok)
            returnValue = msgBox.exec()
            return 0
        
        remarks = surfaceAreaCalculation_remarks.toPlainText().split(';')
        for i in range(1, len(files_T1)):
            language = "english"
            
            driver = Driver_surfaceAreaCalculation_oneRelaxationTime_createFile_importExcel()
            driver.runDriver(surfaceAreaCalculation_materialName.currentText(), Relaxation, surfaceAreaCalculation_bulkName.currentText(), surfaceAreaCalculation_user.toPlainText(), language, remarks[i-1], surfaceAreaCalculation_temperature.toPlainText(), float(surfaceAreaCalculation_surfaceArea_Argon.toPlainText()), float(surfaceAreaCalculation_densityBulk.toPlainText()), float(surfaceAreaCalculation_particleDensity.toPlainText()), surfaceAreaCalculation_dateTime.dateTime().toString("yyyyMMdd"), numOfConcentrations, files_T1, files_T2, filespath_T1, filespath_T2, liquidmassfromTable[i], particlemassfromTable[i], self.referenceFilepath, self.materialName_ReferenceMesurementFiles, self.date_ReferenceMesurementFiles, i)



    def show(self):
        self.main_win.show()
        
        
if __name__ == '__main__':
    app = QtWidgets.QApplication(sys.argv)
    main_win = GUI_MainWindow()
    main_win.show()
    sys.exit(app.exec())
